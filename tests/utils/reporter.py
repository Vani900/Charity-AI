"""
utils/reporter.py – Writes E2E test results to a formatted Excel workbook.

Columns:
  Test ID | Module | Flow Step | Test Name | Description |
  Expected Result | Actual Result | Status | Duration (s) | Error/Notes | Screenshot | Timestamp
"""

import os
import time
import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

RESULTS_DIR = Path(__file__).parent.parent / "results"
SCREENSHOTS_DIR = RESULTS_DIR / "screenshots"

RESULTS_DIR.mkdir(exist_ok=True)
SCREENSHOTS_DIR.mkdir(exist_ok=True)

# ─── Colour palette ─────────────────────────────────────────────
C_HEADER_BG   = "1E3A5F"   # Dark navy
C_HEADER_FG   = "FFFFFF"   # White text
C_PASS_BG     = "D6F5D6"   # Light green
C_PASS_FG     = "1A7431"
C_FAIL_BG     = "FADADD"   # Light red
C_FAIL_FG     = "C0392B"
C_SKIP_BG     = "FFF3CD"   # Light amber
C_SKIP_FG     = "856404"
C_ROW_ALT     = "F0F6FF"   # Alternating row tint
C_MODULE_AUTH = "7C3AED"
C_MODULE_DONOR= "2563EB"
C_MODULE_NGO  = "059669"
C_MODULE_ADMIN= "DC2626"

MODULE_COLORS = {
    "Auth":  "EDE9FE",
    "Donor": "DBEAFE",
    "NGO":   "D1FAE5",
    "Admin": "FEE2E2",
}


def _side():
    return Side(border_style="thin", color="D0D7DE")


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


class TestReporter:
    """Collects test results and serialises them into a polished Excel file."""

    COLUMNS = [
        ("Test ID",         10),
        ("Module",          10),
        ("Flow Step",       22),
        ("Test Name",       35),
        ("Description",     48),
        ("Expected Result", 38),
        ("Actual Result",   38),
        ("Status",          10),
        ("Duration (s)",    13),
        ("Error / Notes",   40),
        ("Screenshot",      32),
        ("Timestamp",       22),
    ]

    def __init__(self):
        self._records: list[dict] = []
        self._counter = 1
        self.output_path = RESULTS_DIR / self._filename()

    # ── Public API ──────────────────────────────────────────────
    def record(
        self,
        module: str,
        flow_step: str,
        test_name: str,
        description: str,
        expected: str,
        actual: str,
        status: str,          # "PASS" | "FAIL" | "SKIP"
        duration: float,
        error: str = "",
        screenshot: str = "",
    ):
        self._records.append({
            "Test ID":        f"TC-{self._counter:03d}",
            "Module":         module,
            "Flow Step":      flow_step,
            "Test Name":      test_name,
            "Description":    description,
            "Expected Result":expected,
            "Actual Result":  actual,
            "Status":         status.upper(),
            "Duration (s)":   round(duration, 2),
            "Error / Notes":  error,
            "Screenshot":     screenshot,
            "Timestamp":      datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
        self._counter += 1

    def save(self):
        wb = openpyxl.Workbook()

        self._build_summary_sheet(wb)
        self._build_detail_sheet(wb)

        wb.save(self.output_path)
        print(f"\n✅  Excel report saved → {self.output_path}\n")
        return self.output_path

    # ── Internal helpers ────────────────────────────────────────
    @staticmethod
    def _filename():
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"CharityAI_TestReport_{ts}.xlsx"

    def _build_summary_sheet(self, wb: openpyxl.Workbook):
        ws = wb.active
        ws.title = "📊 Summary"
        ws.sheet_view.showGridLines = False

        total  = len(self._records)
        passed = sum(1 for r in self._records if r["Status"] == "PASS")
        failed = sum(1 for r in self._records if r["Status"] == "FAIL")
        skipped= sum(1 for r in self._records if r["Status"] == "SKIP")
        pct    = round((passed / total * 100) if total else 0, 1)

        # ── Title banner ────────────────────────────────────────
        ws.merge_cells("B2:G2")
        cell = ws["B2"]
        cell.value = "CharityAI — Selenium End-to-End Test Report"
        cell.font = Font(name="Calibri", bold=True, size=18, color=C_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 36

        ws.merge_cells("B3:G3")
        sub = ws["B3"]
        run_time = datetime.datetime.now().strftime("%B %d, %Y  %H:%M")
        sub.value = f"Generated: {run_time}   |   Base URL: http://localhost:5173"
        sub.font = Font(name="Calibri", italic=True, size=11, color="555555")
        sub.fill = PatternFill("solid", fgColor="EEF3FB")
        sub.alignment = Alignment(horizontal="center")

        # ── KPI Boxes ───────────────────────────────────────────
        boxes = [
            ("B", 5, "Total Tests", str(total),  "1E3A5F", "FFFFFF"),
            ("D", 5, "Passed",      str(passed),  "1A7431", "D6F5D6"),
            ("F", 5, "Failed",      str(failed),  "C0392B", "FADADD"),
            ("H", 5, "Pass Rate",   f"{pct}%",    "0D6EFD", "DBEAFE"),
        ]
        for col, row, label, val, fg, bg in boxes:
            ws.merge_cells(f"{col}{row}:{col}{row+1}")
            c = ws[f"{col}{row}"]
            c.value = label
            c.font = Font(bold=True, color=fg, size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="bottom")
            ws.row_dimensions[row].height = 18

            c2 = ws[f"{col}{row+1}"]
            c2.value = val
            c2.font = Font(bold=True, color=fg, size=22)
            c2.fill = PatternFill("solid", fgColor=bg)
            c2.alignment = Alignment(horizontal="center", vertical="top")
            ws.row_dimensions[row+1].height = 28

        # ── Per-module breakdown ─────────────────────────────────
        ws["B8"].value = "Module Breakdown"
        ws["B8"].font  = Font(bold=True, size=13, color=C_HEADER_BG)
        ws.row_dimensions[8].height = 22

        headers = ["Module", "Total", "Pass", "Fail", "Skip", "Pass %"]
        for ci, h in enumerate(headers, start=2):
            c = ws.cell(row=9, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
            c.alignment = Alignment(horizontal="center")
            c.border = _border()
            ws.column_dimensions[get_column_letter(ci)].width = 14
        ws.row_dimensions[9].height = 20

        modules = sorted(set(r["Module"] for r in self._records))
        for ri, mod in enumerate(modules, start=10):
            recs = [r for r in self._records if r["Module"] == mod]
            p = sum(1 for r in recs if r["Status"] == "PASS")
            f = sum(1 for r in recs if r["Status"] == "FAIL")
            s = sum(1 for r in recs if r["Status"] == "SKIP")
            t = len(recs)
            mp = round((p / t * 100) if t else 0, 1)
            bg = MODULE_COLORS.get(mod, "FFFFFF")
            row_vals = [mod, t, p, f, s, f"{mp}%"]
            for ci, v in enumerate(row_vals, start=2):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="center")
                c.border = _border()
                c.font = Font(size=10)

    def _build_detail_sheet(self, wb: openpyxl.Workbook):
        ws = wb.create_sheet("📋 Test Details")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        # ── Header row ──────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor=C_HEADER_BG)
        for ci, (col_name, col_w) in enumerate(self.COLUMNS, start=1):
            c = ws.cell(row=1, column=ci, value=col_name)
            c.font = Font(bold=True, color=C_HEADER_FG, size=10, name="Calibri")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _border()
            ws.column_dimensions[get_column_letter(ci)].width = col_w
        ws.row_dimensions[1].height = 24

        # ── Data rows ───────────────────────────────────────────
        for ri, record in enumerate(self._records, start=2):
            status = record["Status"]
            if status == "PASS":
                row_bg = C_PASS_BG if ri % 2 == 0 else "FFFFFF"
            elif status == "FAIL":
                row_bg = C_FAIL_BG
            else:
                row_bg = C_SKIP_BG

            row_fill = PatternFill("solid", fgColor=row_bg)

            for ci, (col_name, _) in enumerate(self.COLUMNS, start=1):
                val = record.get(col_name, "")
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = row_fill
                c.border = _border()
                c.font = Font(size=9, name="Calibri")
                c.alignment = Alignment(
                    horizontal="center" if col_name in ("Test ID", "Module", "Status", "Duration (s)") else "left",
                    vertical="center",
                    wrap_text=True,
                )

                # Status cell colouring
                if col_name == "Status":
                    if status == "PASS":
                        c.font = Font(bold=True, color=C_PASS_FG, size=10)
                    elif status == "FAIL":
                        c.font = Font(bold=True, color=C_FAIL_FG, size=10)
                    else:
                        c.font = Font(bold=True, color=C_SKIP_FG, size=10)

            ws.row_dimensions[ri].height = 30


def take_screenshot(driver, name: str) -> str:
    """Take a screenshot and return the file path."""
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = SCREENSHOTS_DIR / f"{name}_{ts}.png"
    driver.save_screenshot(str(fname))
    return str(fname)
